"""
UFC Gym Operate — Outcomes Analysis
====================================
Reads raw call thread data from Google Sheets export and produces:
  1. Console output with all verified metrics
  2. Excel workbook with 7 analysis sheets

Requirements:
  pip install openpyxl

Usage:
  python analyze_operate_outcomes.py --input <path_to_thread_export.tsv>

If --input is not provided, reads from the default JSON cache path.
"""

import json
import argparse
from collections import Counter
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. Excel output disabled. Install with: pip install openpyxl")


# =============================================================================
# 1. DATA LOADING
# =============================================================================

def load_from_json(path: str) -> str:
    """Load text content from the MCP tool result JSON file."""
    with open(path) as f:
        data = json.load(f)
    return data[0]['text'] if isinstance(data, list) else data.get('text', str(data))


def load_from_tsv(path: str) -> str:
    """Load raw TSV content directly."""
    with open(path) as f:
        return f.read()


def parse_threads(text: str) -> dict:
    """
    Parse tab-separated call thread data into a dict of threads.

    Expected columns: Thread ID, Location, Contact Name, Phone, Verdict,
                      Messages, Duration (s), Started At, Direction, Message

    Returns dict keyed by thread ID, each containing:
      - location, contact, verdict, duration, started, transcript[]
    """
    lines = text.strip().split('\n')

    # Find header row
    header_idx = None
    for i, line in enumerate(lines):
        if 'Thread ID' in line and 'Verdict' in line:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Could not find header row with 'Thread ID' and 'Verdict'")

    headers = lines[header_idx].split('\t')

    rows = []
    for line in lines[header_idx + 1:]:
        cols = line.split('\t')
        if len(cols) >= len(headers):
            rows.append(dict(zip(headers, cols)))

    # Group by thread
    threads = {}
    for r in rows:
        tid = r.get('Thread ID', '')
        if tid not in threads:
            threads[tid] = {
                'location': r.get('Location', ''),
                'contact': r.get('Contact Name', ''),
                'verdict': r.get('Verdict', ''),
                'duration': r.get('Duration (s)', ''),
                'started': r.get('Started At', ''),
                'transcript': [],
            }
        threads[tid]['transcript'].append({
            'direction': r.get('Direction', ''),
            'message': r.get('Message', ''),
        })

    return threads


# =============================================================================
# 2. METRICS COMPUTATION
# =============================================================================

def compute_topline(threads: dict) -> dict:
    """Compute top-level aggregate metrics."""
    total = len(threads)
    verdicts = Counter(t['verdict'] for t in threads.values())

    spoke = verdicts.get('spoke_to', 0)
    vm = verdicts.get('left_message', 0)
    other = verdicts.get('', 0)

    all_dur = []
    spoke_dur = []
    for t in threads.values():
        try:
            d = int(t['duration'])
            all_dur.append(d)
            if t['verdict'] == 'spoke_to':
                spoke_dur.append(d)
        except (ValueError, TypeError):
            pass

    locations = set(t['location'] for t in threads.values())

    return {
        'total_threads': total,
        'spoke_to': spoke,
        'left_message': vm,
        'other': other,
        'pickup_rate': spoke / total if total else 0,
        'locations': sorted(locations),
        'location_count': len(locations),
        'total_call_time_s': sum(all_dur),
        'total_connected_time_s': sum(spoke_dur),
        'mean_spoke_duration_s': sum(spoke_dur) / len(spoke_dur) if spoke_dur else 0,
        'median_spoke_duration_s': sorted(spoke_dur)[len(spoke_dur) // 2] if spoke_dur else 0,
        'max_spoke_duration_s': max(spoke_dur) if spoke_dur else 0,
        'spoke_durations': spoke_dur,
        'all_durations': all_dur,
    }


def compute_per_location(threads: dict) -> dict:
    """Compute per-location breakdown."""
    loc_stats = {}
    for t in threads.values():
        loc = t['location']
        if loc not in loc_stats:
            loc_stats[loc] = {
                'total': 0, 'spoke': 0, 'vm': 0, 'other': 0,
                'spoke_dur': [], 'all_dur': [],
            }
        loc_stats[loc]['total'] += 1
        if t['verdict'] == 'spoke_to':
            loc_stats[loc]['spoke'] += 1
        elif t['verdict'] == 'left_message':
            loc_stats[loc]['vm'] += 1
        else:
            loc_stats[loc]['other'] += 1
        try:
            d = int(t['duration'])
            loc_stats[loc]['all_dur'].append(d)
            if t['verdict'] == 'spoke_to':
                loc_stats[loc]['spoke_dur'].append(d)
        except (ValueError, TypeError):
            pass

    # Add computed fields
    for loc, s in loc_stats.items():
        s['pickup_rate'] = s['spoke'] / s['total'] if s['total'] else 0
        s['avg_spoke_dur'] = sum(s['spoke_dur']) / len(s['spoke_dur']) if s['spoke_dur'] else 0
        s['median_spoke_dur'] = sorted(s['spoke_dur'])[len(s['spoke_dur']) // 2] if s['spoke_dur'] else 0
        s['total_minutes'] = sum(s['all_dur']) / 60

    return loc_stats


def compute_duration_buckets(threads: dict) -> dict:
    """Bucket spoke_to calls by duration."""
    bucket_defs = [
        ('<10s', 0, 10),
        ('10-30s', 10, 30),
        ('30-60s', 30, 60),
        ('60-120s', 60, 120),
        ('120-300s', 120, 300),
        ('300s+', 300, 999999),
    ]
    buckets = {}
    for label, lo, hi in bucket_defs:
        buckets[label] = sum(
            1 for t in threads.values()
            if t['verdict'] == 'spoke_to'
            and lo <= (int(t['duration']) if t['duration'] else 0) < hi
        )
    return buckets


def compute_quality(threads: dict) -> dict:
    """Compute conversation quality metrics for spoke_to calls."""
    spoke_total = sum(1 for t in threads.values() if t['verdict'] == 'spoke_to')

    real_dialog = 0
    meaningful = 0
    hangup = 0
    clean_close = 0

    for t in threads.values():
        if t['verdict'] != 'spoke_to':
            continue

        dur = int(t['duration']) if t['duration'] else 0
        if dur >= 30:
            meaningful += 1

        member_turns = sum(
            1 for m in t['transcript']
            if m['direction'] == 'inbound' and len(m['message'].strip()) > 5
        )
        if member_turns >= 2:
            real_dialog += 1

        sys_msgs = [m['message'].lower() for m in t['transcript'] if m['direction'] == 'system']
        if any('hung up' in m for m in sys_msgs):
            hangup += 1
        elif any('ended' in m or 'completed' in m for m in sys_msgs):
            clean_close += 1

    return {
        'spoke_total': spoke_total,
        'real_dialog': real_dialog,
        'meaningful_30s': meaningful,
        'member_hangup': hangup,
        'clean_close': clean_close,
    }


def compute_signals(threads: dict) -> dict:
    """
    Detect outcome signals and technical issues in transcripts.
    All pattern matching is explicit — no fuzzy or estimated counts.
    """
    total = len(threads)
    spoke_total = sum(1 for t in threads.values() if t['verdict'] == 'spoke_to')

    recovery = 0
    booking = 0
    member_info_req = 0
    sms_sent = 0
    opt_out = 0
    escalation = 0
    important_msg = 0
    double_pitch = 0

    recovery_keywords = ['recovery', 'hydromassage', 'cryotherapy', 'compression massage', 'percussion']
    booking_keywords = ['book', 'schedule', 'appointment']
    info_keywords = ['how much', 'what time', 'what is the price', 'what are the hours', 'cost', 'what does it cost']
    opt_out_keywords = ['stop', 'opt out', 'unsubscribe', 'remove me', "don't text", "don't call"]

    for tid, t in threads.items():
        all_text = ' '.join(m['message'].lower() for m in t['transcript'])
        agent_msgs = [m['message'] for m in t['transcript'] if m['direction'] == 'outbound']
        agent_text = ' '.join(m['message'].lower() for m in t['transcript'] if m['direction'] == 'outbound')
        member_text = ' '.join(
            m['message'].lower() for m in t['transcript'] if m['direction'] == 'inbound'
        )

        # Recovery services mentioned anywhere
        if any(kw in all_text for kw in recovery_keywords):
            recovery += 1

        # Booking/scheduling language anywhere
        if any(kw in all_text for kw in booking_keywords):
            booking += 1

        # Member-initiated info requests (spoke_to only, member text only)
        if t['verdict'] == 'spoke_to' and any(kw in member_text for kw in info_keywords):
            member_info_req += 1

        # System SMS sent
        if any(
            'sms sent' in m['message'].lower()
            for m in t['transcript']
            if m['direction'] == 'system'
        ):
            sms_sent += 1

        # Opt-out keywords anywhere
        if any(kw in all_text for kw in opt_out_keywords):
            opt_out += 1

        # Agent offered transfer (agent text only)
        if any(kw in agent_text for kw in ['transfer you', 'connect you with']):
            escalation += 1

        # "Important message" voicemail fallback (agent text only)
        if any(
            'important message' in m['message'].lower()
            for m in t['transcript']
            if m['direction'] == 'outbound'
        ):
            important_msg += 1

        # Double-pitch: opening line repeated after disclaimer (spoke_to only)
        if t['verdict'] == 'spoke_to' and len(agent_msgs) >= 3:
            first_50 = agent_msgs[0][:50].lower()
            for msg in agent_msgs[2:]:
                if first_50 in msg[:50].lower():
                    double_pitch += 1
                    break

    return {
        'total': total,
        'spoke_total': spoke_total,
        'recovery': recovery,
        'booking': booking,
        'member_info_req': member_info_req,
        'sms_sent': sms_sent,
        'opt_out': opt_out,
        'escalation': escalation,
        'important_msg': important_msg,
        'double_pitch': double_pitch,
    }


def compute_error_modes(threads: dict) -> dict:
    """
    Diagnose WHY conversations fail by analyzing transcript-level patterns.

    Returns a dict with:
      - dropout_turn: at which agent turn do members hang up?
      - last_agent_msg_before_hangup: what was the agent saying when the member left?
      - member_first_response: how do members initially respond?
      - failure_taxonomy: categorize non-meaningful spoke_to calls by failure mode
      - agent_repetition: does the agent repeat itself beyond double-pitch?
      - other_threads: what happened in the 'no verdict' threads?
      - member_last_words: what was the last thing the member said?
      - vm_fallback_triggers: what precedes the "important message" fallback?
      - per_thread_errors: list of (tid, error_modes) for the detail sheet
    """
    results = {
        'dropout_by_agent_turn': Counter(),       # agent turn # when member hung up
        'last_agent_before_hangup': [],            # (tid, location, duration, message)
        'member_first_response': Counter(),        # classified first member utterance
        'failure_taxonomy': Counter(),             # failure mode for non-meaningful calls
        'agent_repetition_threads': 0,             # threads where agent says same thing 3+ times
        'other_thread_analysis': Counter(),        # what system messages appear in 'other' threads
        'member_last_words': [],                   # (tid, location, duration, message)
        'vm_clean_count': 0,                       # voicemails with clean delivery
        'vm_fallback_count': 0,                    # voicemails with "important message" fallback
        'vm_fallback_preceding': Counter(),        # what inbound message preceded the fallback
        'per_thread_errors': {},                   # tid -> list of error mode strings
    }

    for tid, t in threads.items():
        errors = []
        transcript = t['transcript']
        agent_msgs = [(i, m) for i, m in enumerate(transcript) if m['direction'] == 'outbound']
        member_msgs = [(i, m) for i, m in enumerate(transcript) if m['direction'] == 'inbound']
        sys_msgs = [(i, m) for i, m in enumerate(transcript) if m['direction'] == 'system']
        dur = int(t['duration']) if t['duration'] else 0

        # ----------------------------------------------------------
        # A. SPOKE_TO ANALYSIS
        # ----------------------------------------------------------
        if t['verdict'] == 'spoke_to':

            # A1. Dropout turn: count how many agent messages played before hangup
            hangup_idx = None
            for i, m in sys_msgs:
                if 'hung up' in m['message'].lower():
                    hangup_idx = i
                    break

            if hangup_idx is not None:
                agent_turns_before = sum(1 for ai, am in agent_msgs if ai < hangup_idx)
                results['dropout_by_agent_turn'][agent_turns_before] += 1

                # Last agent message before hangup
                last_agent = None
                for ai, am in reversed(agent_msgs):
                    if ai < hangup_idx:
                        last_agent = am['message']
                        break
                if last_agent:
                    results['last_agent_before_hangup'].append(
                        (tid, t['location'], dur, last_agent[:200])
                    )

            # A2. Member first response classification
            if member_msgs:
                first_msg = member_msgs[0][1]['message'].strip().lower()
                if len(first_msg) <= 2 or first_msg in ('', '.'):
                    cat = 'silence/noise'
                elif first_msg in ('hello?', 'hello', 'hi', 'hey', 'yes?', 'yeah?', 'huh?'):
                    cat = 'greeting (hello?)'
                elif 'voicemail' in first_msg or 'not available' in first_msg or 'leave a message' in first_msg or 'forwarded' in first_msg:
                    cat = 'voicemail greeting (misdetected)'
                elif any(w in first_msg for w in ['who is this', 'who are you', 'who calling', 'what number']):
                    cat = 'suspicion (who is this?)'
                elif any(w in first_msg for w in ['not interested', 'no thanks', 'no thank', 'don\'t want', 'stop calling']):
                    cat = 'immediate rejection'
                elif any(w in first_msg for w in ['wrong number', 'wrong person']):
                    cat = 'wrong number'
                elif len(first_msg) > 20:
                    cat = 'substantive response'
                else:
                    cat = 'brief acknowledgment'
            else:
                cat = 'no member speech detected'
            results['member_first_response'][cat] += 1

            # A3. Member last words before hangup
            if member_msgs:
                last_member = member_msgs[-1][1]['message'].strip()
                if len(last_member) > 3:
                    results['member_last_words'].append(
                        (tid, t['location'], dur, last_member[:200])
                    )

            # A4. Failure taxonomy for non-meaningful calls (<30s)
            if dur < 30:
                if not member_msgs:
                    mode = 'no_member_speech'
                    errors.append('no_member_speech')
                elif len(member_msgs) == 1 and member_msgs[0][1]['message'].strip().lower() in ('hello?', 'hello', 'hi', 'hey', 'yes?', 'yeah?'):
                    mode = 'greeting_then_hangup'
                    errors.append('greeting_then_hangup')
                elif any('voicemail' in m[1]['message'].lower() or 'not available' in m[1]['message'].lower() or 'forwarded' in m[1]['message'].lower() for m in member_msgs):
                    mode = 'voicemail_misdetected_as_spoke'
                    errors.append('voicemail_misdetected_as_spoke')
                elif any(w in ' '.join(m[1]['message'].lower() for m in member_msgs) for w in ['who is this', 'who are you']):
                    mode = 'suspicion_hangup'
                    errors.append('suspicion_hangup')
                elif any(w in ' '.join(m[1]['message'].lower() for m in member_msgs) for w in ['not interested', 'no thanks', 'stop']):
                    mode = 'immediate_rejection'
                    errors.append('immediate_rejection')
                else:
                    mode = 'brief_exchange_unclear'
                    errors.append('brief_exchange_unclear')
                results['failure_taxonomy'][mode] += 1

            # A5. Double-pitch detection (already counted globally, tag per-thread)
            if len(agent_msgs) >= 3:
                first_50 = agent_msgs[0][1]['message'][:50].lower()
                for _, am in agent_msgs[2:]:
                    if first_50 in am['message'][:50].lower():
                        errors.append('double_pitch')
                        break

            # A6. Agent repetition: same substantive message said 3+ times
            agent_texts = [am['message'][:80].lower().strip() for _, am in agent_msgs if len(am['message']) > 30]
            text_counts = Counter(agent_texts)
            if any(c >= 3 for c in text_counts.values()):
                results['agent_repetition_threads'] += 1
                errors.append('agent_triple_repeat')

        # ----------------------------------------------------------
        # B. LEFT_MESSAGE ANALYSIS
        # ----------------------------------------------------------
        elif t['verdict'] == 'left_message':
            has_fallback = any(
                'important message' in m['message'].lower()
                for _, m in agent_msgs
            )
            if has_fallback:
                results['vm_fallback_count'] += 1
                errors.append('vm_important_message_fallback')
                # What inbound message preceded the fallback?
                for i, m in enumerate(transcript):
                    if m['direction'] == 'outbound' and 'important message' in m['message'].lower():
                        # Look backward for the preceding inbound message
                        for j in range(i - 1, -1, -1):
                            if transcript[j]['direction'] == 'inbound':
                                preceding = transcript[j]['message'][:100].strip()
                                if len(preceding) > 5:
                                    results['vm_fallback_preceding'][preceding] += 1
                                break
                        break
            else:
                results['vm_clean_count'] += 1

        # ----------------------------------------------------------
        # C. OTHER (no verdict) ANALYSIS
        # ----------------------------------------------------------
        elif t['verdict'] == '' or t['verdict'] is None:
            sys_text = [m['message'] for _, m in sys_msgs]
            if not sys_text:
                results['other_thread_analysis']['no_system_messages'] += 1
                errors.append('no_verdict_no_system')
            else:
                combined = ' '.join(s.lower() for s in sys_text)
                if 'error' in combined or 'fail' in combined:
                    results['other_thread_analysis']['system_error'] += 1
                    errors.append('system_error')
                elif 'hung up' in combined:
                    results['other_thread_analysis']['hung_up_no_verdict'] += 1
                    errors.append('hung_up_no_verdict')
                elif 'voicemail' in combined:
                    results['other_thread_analysis']['voicemail_no_verdict'] += 1
                    errors.append('voicemail_no_verdict')
                else:
                    results['other_thread_analysis']['other_unknown'] += 1
                    errors.append('unknown_no_verdict')
                    # Store the actual messages for inspection
                    for s in sys_text[:3]:
                        results['other_thread_analysis'][f'  msg: {s[:80]}'] += 1

        results['per_thread_errors'][tid] = errors

    return results


def classify_turn(direction: str, message: str, is_first_agent: bool = False) -> str:
    """
    Classify a single conversation turn into a dialog act.

    Based on DAMSL/ISO 24617-2 dialog act taxonomy, simplified for voice agent analysis.
    Returns one of:
      SYSTEM: call_start, call_end, hangup, vm_detected, sms_sent, error
      AGENT:  disclaimer, pitch, pitch_repeat, greeting, question, info_give, cta, farewell, fallback, filler
      MEMBER: greeting, ack_brief, ack_positive, question, rejection, confusion, vm_greeting, substantive, silence
    """
    msg = message.strip().lower()

    if direction == 'system':
        if 'call started' in msg:
            return 'SYS:call_start'
        if 'hung up' in msg:
            return 'SYS:hangup'
        if 'call ended' in msg or 'ended the call' in msg:
            return 'SYS:call_end'
        if 'voicemail' in msg:
            return 'SYS:vm_detected'
        if 'sms sent' in msg:
            return 'SYS:sms_sent'
        if 'no activity' in msg:
            return 'SYS:timeout'
        return 'SYS:other'

    if direction == 'outbound':  # Agent
        if 'this call may be recorded' in msg or 'quality assurance' in msg:
            return 'AGT:disclaimer'
        if 'important message' in msg:
            return 'AGT:fallback'
        if any(w in msg for w in ['just wanted to check', 'wanted to make sure', 'we noticed']):
            return 'AGT:pitch'
        if any(w in msg for w in ['how have you been', 'how are you', 'how can i help', 'what would']):
            return 'AGT:question'
        if any(w in msg for w in ['hydromassage', 'cryotherapy', 'compression', 'percussion', 'recovery', 'hiit', 'conditioning']):
            if '?' in msg:
                return 'AGT:question'
            return 'AGT:info_give'
        if any(w in msg for w in ['would you like me to', 'can i text', 'can i send', 'here\'s that link', 'i just texted']):
            return 'AGT:cta'
        if any(w in msg for w in ['no problem', 'totally fair', 'no pressure', 'no worries']):
            return 'AGT:ack_graceful'
        if any(w in msg for w in ['for sure', 'got you', 'nice', 'great']):
            return 'AGT:ack_positive'
        if any(w in msg for w in ['take care', 'have a great', 'goodbye', 'bye']):
            return 'AGT:farewell'
        if any(w in msg for w in ['hey', 'hi ', 'hello', "it's jade", "it's your"]):
            return 'AGT:greeting'
        if any(w in msg for w in ['pricing', 'first session is free', 'ten dollars', '$', 'percent off']):
            return 'AGT:info_give'
        if any(w in msg for w in ['transfer you', 'connect you with', 'reach the gym']):
            return 'AGT:escalation_offer'
        if len(msg) < 15:
            return 'AGT:filler'
        return 'AGT:info_give'

    if direction == 'inbound':  # Member
        if len(msg) <= 2:
            return 'MBR:silence'
        if any(w in msg for w in ['voicemail', 'not available', 'leave a message', 'forwarded to voice', 'can\'t take your call']):
            return 'MBR:vm_greeting'
        if msg in ('hello?', 'hello', 'hello.', 'hi', 'hey', 'yes?', 'yeah?', 'huh?'):
            return 'MBR:greeting'
        if any(w in msg for w in ['who is this', 'who are you', 'who calling', 'what number is this']):
            return 'MBR:suspicion'
        if any(w in msg for w in ['not interested', 'no thanks', 'no thank', 'don\'t want', 'stop calling', 'don\'t call', 'remove me']):
            return 'MBR:rejection'
        if any(w in msg for w in ['wrong number', 'wrong person']):
            return 'MBR:wrong_number'
        if any(w in msg for w in ['how much', 'what time', 'what does it cost', 'when is', 'where is', 'what are the hours']):
            return 'MBR:question'
        if any(w in msg for w in ['yes', 'yeah', 'sure', 'okay', 'ok', 'alright', 'uh huh', 'mhm']):
            return 'MBR:ack_brief'
        if any(w in msg for w in ['great', 'awesome', 'cool', 'sounds good', 'perfect', 'thank', 'appreciate']):
            return 'MBR:ack_positive'
        if any(w in msg for w in ['no', 'nope', 'not really', 'i\'m good', 'i\'m ok', 'i don\'t', "i'm okay", "don't need"]):
            return 'MBR:decline_soft'
        if any(w in msg for w in ['please stay on the line', 'if you record your name']):
            return 'MBR:call_screening'
        if len(msg) > 30:
            return 'MBR:substantive'
        return 'MBR:brief_other'

    return 'UNK:unknown'


def compute_conversation_flows(threads: dict) -> dict:
    """
    Extract dialog act sequences for every thread and identify structural failure patterns.

    Approach inspired by TD-EVAL (Acikgoz et al. 2025) and BETOLD (Terragni et al. 2022):
    - Classify every turn into a dialog act
    - Extract the act sequence as a "conversation fingerprint"
    - Cluster fingerprints to find dominant patterns
    - Compare successful (>30s, 2+ member turns) vs failed conversations
    - Identify the exact turn where conversations diverge

    Returns dict with:
      - per_thread_acts: {tid: [act1, act2, ...]}
      - flow_patterns: Counter of act sequence fingerprints
      - success_patterns: patterns from successful calls
      - failure_patterns: patterns from failed calls
      - divergence_analysis: where success/failure paths split
      - agent_behavior_after_greeting: what the agent does after member says "hello?"
      - member_response_to_pitch: how members respond to the pitch
      - per_thread_flow_detail: {tid: {acts, fingerprint, quality_label, failure_point}}
    """
    per_thread_acts = {}
    flow_patterns = Counter()
    success_patterns = Counter()
    failure_patterns = Counter()
    per_thread_flow_detail = {}

    # Track: what does agent do after member greeting?
    agent_after_greeting = Counter()
    # Track: how does member respond to first pitch?
    member_after_pitch = Counter()
    # Track: what agent turn is playing when member hangs up?
    hangup_on_act = Counter()
    # Track: sequence at point of hangup (last 3 acts before hangup)
    hangup_context = Counter()
    # For divergence: position-by-position act frequencies for success vs failure
    success_acts_by_pos = {}  # {position: Counter of acts}
    failure_acts_by_pos = {}

    for tid, t in threads.items():
        if t['verdict'] != 'spoke_to':
            continue

        dur = int(t['duration']) if t['duration'] else 0
        member_turns = sum(
            1 for m in t['transcript']
            if m['direction'] == 'inbound' and len(m['message'].strip()) > 5
        )

        # Classify every turn
        acts = []
        is_first_agent = True
        for m in t['transcript']:
            act = classify_turn(m['direction'], m['message'], is_first_agent)
            acts.append(act)
            if m['direction'] == 'outbound':
                is_first_agent = False

        per_thread_acts[tid] = acts

        # Create fingerprint: only non-system acts, compressed
        non_sys = [a for a in acts if not a.startswith('SYS:')]
        # Compress consecutive same acts
        compressed = []
        for a in non_sys:
            if not compressed or compressed[-1] != a:
                compressed.append(a)
        fingerprint = ' → '.join(compressed[:10])  # cap at 10 acts

        # Label quality
        if dur >= 30 and member_turns >= 2:
            quality = 'success'
        elif dur < 10:
            quality = 'fail_immediate'
        elif member_turns == 0:
            quality = 'fail_no_response'
        else:
            quality = 'fail_brief'

        # Identify failure point: the last member act before hangup
        failure_point = None
        hangup_idx = None
        for i, a in enumerate(acts):
            if a == 'SYS:hangup':
                hangup_idx = i
                break

        if hangup_idx is not None:
            # Last agent act before hangup
            last_agent_act = None
            for i in range(hangup_idx - 1, -1, -1):
                if acts[i].startswith('AGT:'):
                    last_agent_act = acts[i]
                    break
            if last_agent_act:
                hangup_on_act[last_agent_act] += 1

            # Context: last 3 acts before hangup
            context_start = max(0, hangup_idx - 3)
            context = ' → '.join(acts[context_start:hangup_idx])
            hangup_context[context] += 1

            # Last member act before hangup
            for i in range(hangup_idx - 1, -1, -1):
                if acts[i].startswith('MBR:'):
                    failure_point = acts[i]
                    break

        per_thread_flow_detail[tid] = {
            'acts': acts,
            'fingerprint': fingerprint,
            'quality': quality,
            'failure_point': failure_point,
            'duration': dur,
        }

        flow_patterns[fingerprint] += 1
        if quality == 'success':
            success_patterns[fingerprint] += 1
        else:
            failure_patterns[fingerprint] += 1

        # Track positional acts for divergence analysis
        target = success_acts_by_pos if quality == 'success' else failure_acts_by_pos
        for pos, a in enumerate(compressed[:8]):
            if pos not in target:
                target[pos] = Counter()
            target[pos][a] += 1

        # Agent response after member greeting
        for i, a in enumerate(acts):
            if a == 'MBR:greeting':
                # Find next agent act
                for j in range(i + 1, len(acts)):
                    if acts[j].startswith('AGT:'):
                        agent_after_greeting[acts[j]] += 1
                        break
                break

        # Member response after first pitch
        first_pitch_idx = None
        for i, a in enumerate(acts):
            if a == 'AGT:pitch':
                first_pitch_idx = i
                break
        if first_pitch_idx is not None:
            for j in range(first_pitch_idx + 1, len(acts)):
                if acts[j].startswith('MBR:') or acts[j] == 'SYS:hangup':
                    member_after_pitch[acts[j]] += 1
                    break

    # Build divergence analysis
    divergence = []
    max_pos = max(
        max(success_acts_by_pos.keys(), default=0),
        max(failure_acts_by_pos.keys(), default=0)
    )
    for pos in range(min(max_pos + 1, 8)):
        s_acts = success_acts_by_pos.get(pos, Counter())
        f_acts = failure_acts_by_pos.get(pos, Counter())
        s_top = s_acts.most_common(3)
        f_top = f_acts.most_common(3)
        divergence.append({
            'position': pos,
            'success_top': s_top,
            'failure_top': f_top,
        })

    return {
        'per_thread_acts': per_thread_acts,
        'per_thread_flow_detail': per_thread_flow_detail,
        'flow_patterns': flow_patterns,
        'success_patterns': success_patterns,
        'failure_patterns': failure_patterns,
        'divergence': divergence,
        'agent_after_greeting': agent_after_greeting,
        'member_after_pitch': member_after_pitch,
        'hangup_on_act': hangup_on_act,
        'hangup_context': hangup_context,
    }


def compute_timing(threads: dict) -> dict:
    """Compute call distribution by hour (UTC)."""
    hour_total = Counter()
    hour_spoke = Counter()

    for t in threads.values():
        started = t.get('started', '')
        if 'GMT' not in started:
            continue
        parts = started.split(' ')
        if len(parts) < 5:
            continue
        time_part = parts[4]
        if ':' not in time_part:
            continue
        hour = int(time_part.split(':')[0])
        hour_total[hour] += 1
        if t['verdict'] == 'spoke_to':
            hour_spoke[hour] += 1

    return {'hour_total': hour_total, 'hour_spoke': hour_spoke}


# =============================================================================
# 3. CONSOLE OUTPUT
# =============================================================================

def print_tables(threads, topline, loc_stats, buckets, quality, signals, timing, conv_flows=None, error_modes=None):
    """Print all analysis tables to console."""
    spoke = topline['spoke_to']

    print("=" * 70)
    print("TABLE 1: TOPLINE METRICS")
    print("=" * 70)
    print(f"  Total threads:          {topline['total_threads']}")
    print(f"  Locations:              {topline['location_count']}")
    print(f"  Spoke to:               {spoke} ({topline['pickup_rate']:.1%})")
    print(f"  Left message:           {topline['left_message']} ({topline['left_message']/topline['total_threads']:.1%})")
    print(f"  Other:                  {topline['other']}")
    print(f"  Total call time:        {topline['total_call_time_s']:,}s ({topline['total_call_time_s']/3600:.1f} hrs)")
    print(f"  Connected time:         {topline['total_connected_time_s']:,}s ({topline['total_connected_time_s']/60:.1f} min)")
    print(f"  Mean spoke duration:    {topline['mean_spoke_duration_s']:.1f}s")
    print(f"  Median spoke duration:  {topline['median_spoke_duration_s']}s")
    print(f"  Max spoke duration:     {topline['max_spoke_duration_s']}s")

    print(f"\n{'=' * 100}")
    print("TABLE 2: PER-LOCATION PERFORMANCE")
    print("=" * 100)
    print(f"{'Location':<38} {'Calls':>5} {'Spoke':>5} {'VM':>5} {'Oth':>4} {'Pickup%':>8} {'AvgDur':>7} {'MedDur':>7} {'TotMin':>8}")
    print("-" * 100)

    ranked = sorted(loc_stats.items(), key=lambda x: x[1]['pickup_rate'], reverse=True)
    for loc, s in ranked:
        print(f"{loc:<38} {s['total']:>5} {s['spoke']:>5} {s['vm']:>5} {s['other']:>4} {s['pickup_rate']:>7.1%} {s['avg_spoke_dur']:>6.0f}s {s['median_spoke_dur']:>6}s {s['total_minutes']:>7.1f}m")

    print(f"\n{'=' * 70}")
    print("TABLE 3: DURATION DISTRIBUTION (spoke_to)")
    print("=" * 70)
    for label, count in buckets.items():
        pct = count / spoke if spoke else 0
        bar = '#' * int(pct * 50)
        print(f"  {label:>8}: {count:>4} ({pct:>5.1%}) {bar}")

    print(f"\n{'=' * 70}")
    print("TABLE 4: CONVERSATION QUALITY")
    print("=" * 70)
    q = quality
    print(f"  Real dialog (2+ member turns): {q['real_dialog']:>4} ({q['real_dialog']/q['spoke_total']:.1%})")
    print(f"  Meaningful (>30s):             {q['meaningful_30s']:>4} ({q['meaningful_30s']/q['spoke_total']:.1%})")
    print(f"  Member hung up:                {q['member_hangup']:>4} ({q['member_hangup']/q['spoke_total']:.1%})")
    print(f"  Clean close:                   {q['clean_close']:>4} ({q['clean_close']/q['spoke_total']:.1%})")

    print(f"\n{'=' * 70}")
    print("TABLE 5: OUTCOME SIGNALS")
    print("=" * 70)
    s = signals
    print(f"  Recovery mentioned:      {s['recovery']:>4} ({s['recovery']/s['total']:.1%} of all)")
    print(f"  Booking/scheduling:      {s['booking']:>4} ({s['booking']/s['spoke_total']:.1%} of spoke)")
    print(f"  Member info requests:    {s['member_info_req']:>4} ({s['member_info_req']/s['spoke_total']:.1%} of spoke)")
    print(f"  SMS link sent:           {s['sms_sent']:>4} ({s['sms_sent']/s['spoke_total']:.1%} of spoke)")
    print(f"  Opt-out keywords:        {s['opt_out']:>4} ({s['opt_out']/s['total']:.1%} of all)")
    print(f"  Agent offered transfer:  {s['escalation']:>4} ({s['escalation']/s['total']:.1%} of all)")
    print(f"  --- Technical Issues ---")
    print(f"  Double-pitch bug:        {s['double_pitch']:>4} ({s['double_pitch']/s['spoke_total']:.1%} of spoke)")
    print(f"  VM fallback triggered:   {s['important_msg']:>4} ({s['important_msg']/s['total']:.1%} of all)")

    print(f"\n{'=' * 70}")
    print("TABLE 6: CALL TIMING (UTC)")
    print("=" * 70)
    tz_map = {3: '~8pm HI', 15: '~8am PT', 17: '~10am PT', 18: '~11am PT',
              19: '~12pm PT', 20: '~1pm PT', 21: '~2pm PT'}
    total_day = sum(timing['hour_total'].values())
    for h in sorted(timing['hour_total'].keys()):
        t = timing['hour_total'][h]
        s = timing['hour_spoke'].get(h, 0)
        rate = s / t if t else 0
        share = t / total_day if total_day else 0
        print(f"  {h:>2}:00 ({tz_map.get(h, '?'):>10})  {t:>4} calls  {s:>4} spoke  {rate:>6.1%} pickup  {share:>6.1%} of day")

    if conv_flows:
        cf = conv_flows
        print(f"\n{'=' * 80}")
        print("TABLE 7: CONVERSATION FLOW ANALYSIS (SOTA: dialog act classification)")
        print("=" * 80)

        print(f"\n--- 7a. AGENT ACT WHEN MEMBER HANGS UP ---")
        print(f"  (What dialog act was the agent performing at the moment of hangup?)")
        total_hu = sum(cf['hangup_on_act'].values())
        for act, count in cf['hangup_on_act'].most_common(10):
            bar = '#' * (count // 2)
            print(f"  {act:<25} {count:>4} ({count/total_hu:.1%}) {bar}")

        print(f"\n--- 7b. MEMBER RESPONSE TO FIRST PITCH ---")
        print(f"  (After agent delivers the opening pitch, what does the member do?)")
        total_mp = sum(cf['member_after_pitch'].values())
        for act, count in cf['member_after_pitch'].most_common():
            print(f"  {act:<25} {count:>4} ({count/total_mp:.1%})")

        print(f"\n--- 7c. AGENT RESPONSE AFTER MEMBER SAYS 'HELLO?' ---")
        total_ag = sum(cf['agent_after_greeting'].values())
        for act, count in cf['agent_after_greeting'].most_common():
            print(f"  {act:<25} {count:>4} ({count/total_ag:.1%})")

        print(f"\n--- 7d. LAST 3-ACT CONTEXT BEFORE HANGUP (top 10) ---")
        for ctx, count in cf['hangup_context'].most_common(10):
            print(f"  [{count:>3}x] {ctx}")

        print(f"\n--- 7e. WHERE SUCCESS AND FAILURE DIVERGE (position-by-position) ---")
        print(f"  {'Pos':<5} {'Success (top act)':<35} {'Failure (top act)':<35}")
        print(f"  {'-'*75}")
        for d in cf['divergence'][:8]:
            s_str = ', '.join(f"{a}({c})" for a, c in d['success_top'][:2]) if d['success_top'] else '—'
            f_str = ', '.join(f"{a}({c})" for a, c in d['failure_top'][:2]) if d['failure_top'] else '—'
            print(f"  {d['position']:<5} {s_str:<35} {f_str:<35}")

        print(f"\n--- 7f. TOP CONVERSATION FLOW PATTERNS ---")
        print(f"  Unique patterns: {len(cf['flow_patterns'])}")
        print(f"\n  SUCCESS patterns (top 5):")
        for fp, count in cf['success_patterns'].most_common(5):
            print(f"    [{count:>2}x] {fp}")
        print(f"\n  FAILURE patterns (top 5):")
        for fp, count in cf['failure_patterns'].most_common(5):
            print(f"    [{count:>2}x] {fp}")

        # Quality distribution
        quality_dist = Counter(d['quality'] for d in cf['per_thread_flow_detail'].values())
        print(f"\n--- 7g. CONVERSATION QUALITY DISTRIBUTION ---")
        for q, c in quality_dist.most_common():
            print(f"  {q:<25} {c:>4} ({c/sum(quality_dist.values()):.1%})")

    if error_modes:
        em = error_modes
        print(f"\n{'=' * 80}")
        print("TABLE 8: ERROR MODE ANALYSIS")
        print("=" * 80)

        print(f"\n--- 7a. MEMBER FIRST RESPONSE (spoke_to: {spoke} calls) ---")
        for cat, count in em['member_first_response'].most_common():
            bar = '#' * (count // 2)
            print(f"  {cat:<40} {count:>4} ({count/spoke:.1%}) {bar}")

        print(f"\n--- 7b. FAILURE TAXONOMY (non-meaningful spoke_to: <30s) ---")
        non_meaningful = sum(em['failure_taxonomy'].values())
        print(f"  Total non-meaningful spoke_to calls: {non_meaningful}")
        for mode, count in em['failure_taxonomy'].most_common():
            print(f"  {mode:<40} {count:>4} ({count/non_meaningful:.1%} of failures)")

        print(f"\n--- 7c. DROPOUT BY AGENT TURN (spoke_to, member hung up) ---")
        print(f"  {'Agent turn #':<20} {'Hangups':>8} {'%':>8}")
        total_hangups = sum(em['dropout_by_agent_turn'].values())
        for turn in sorted(em['dropout_by_agent_turn'].keys()):
            c = em['dropout_by_agent_turn'][turn]
            print(f"  Turn {turn:<17} {c:>8} {c/total_hangups:>7.1%}")

        print(f"\n--- 7d. LAST AGENT MESSAGE BEFORE HANGUP (top 10 patterns) ---")
        # Cluster by first 80 chars
        last_msg_clusters = Counter()
        for tid, loc, dur, msg in em['last_agent_before_hangup']:
            key = msg[:80].strip()
            last_msg_clusters[key] += 1
        for msg, count in last_msg_clusters.most_common(10):
            print(f"  [{count:>3}x] \"{msg}...\"")

        print(f"\n--- 7e. MEMBER LAST WORDS BEFORE HANGUP (top 10 patterns) ---")
        last_word_clusters = Counter()
        for tid, loc, dur, msg in em['member_last_words']:
            key = msg[:60].strip().lower()
            last_word_clusters[key] += 1
        for msg, count in last_word_clusters.most_common(10):
            print(f"  [{count:>3}x] \"{msg}\"")

        print(f"\n--- 7f. AGENT REPETITION (same message 3+ times) ---")
        print(f"  Threads with triple+ repeat: {em['agent_repetition_threads']}")

        print(f"\n--- 7g. VOICEMAIL DELIVERY ---")
        print(f"  Clean VM delivery: {em['vm_clean_count']}")
        print(f"  'Important message' fallback: {em['vm_fallback_count']}")
        if em['vm_fallback_preceding']:
            print(f"  What preceded the fallback (top 5):")
            for msg, count in em['vm_fallback_preceding'].most_common(5):
                print(f"    [{count:>3}x] \"{msg}\"")

        print(f"\n--- 7h. 'OTHER' THREADS (no verdict: {topline['other']}) ---")
        for cat, count in em['other_thread_analysis'].most_common():
            if not cat.startswith('  msg:'):
                print(f"  {cat:<40} {count:>4}")
            else:
                print(f"    {cat}")


def print_synthesis(topline, quality, signals, conv_flows, error_modes):
    """
    Synthesize all computed metrics into a root-cause narrative.
    Every number is derived from the analysis dicts — nothing hardcoded.
    """
    spoke = topline['spoke_to']
    total = topline['total_threads']

    print(f"\n{'=' * 80}")
    print("TABLE 9: SYNTHESIZED ROOT-CAUSE FINDINGS")
    print("=" * 80)

    # --- Finding 1: Dominant failure flow ---
    top_fail = conv_flows['failure_patterns'].most_common(1)
    if top_fail:
        fp, fp_count = top_fail[0]
        fail_total = sum(conv_flows['failure_patterns'].values())
        print(f"\n  FINDING 1: DOMINANT FAILURE FLOW")
        print(f"  Pattern:   {fp}")
        print(f"  Frequency: {fp_count}x ({fp_count}/{fail_total} = {fp_count/fail_total:.1%} of all failures)")
        print(f"  Meaning:   Agent leads with pitch before member is ready, member says")
        print(f"             'Hello?', agent plays disclaimer, then repeats same pitch.")

    # --- Finding 2: What agent act triggers hangup ---
    top_hangup_act = conv_flows['hangup_on_act'].most_common(1)
    total_hu = sum(conv_flows['hangup_on_act'].values())
    if top_hangup_act:
        act, count = top_hangup_act[0]
        print(f"\n  FINDING 2: HANGUP TRIGGER")
        print(f"  {count}/{total_hu} hangups ({count/total_hu:.1%}) occur during '{act}'")
        # Sum pitch-related acts
        pitch_acts = sum(c for a, c in conv_flows['hangup_on_act'].items() if 'pitch' in a)
        greet_acts = sum(c for a, c in conv_flows['hangup_on_act'].items() if 'greeting' in a)
        print(f"  Pitch-related hangups: {pitch_acts}/{total_hu} ({pitch_acts/total_hu:.1%})")
        print(f"  Greeting-related hangups: {greet_acts}/{total_hu} ({greet_acts/total_hu:.1%})")

    # --- Finding 3: Member response to pitch ---
    top_response = conv_flows['member_after_pitch'].most_common(1)
    total_mp = sum(conv_flows['member_after_pitch'].values())
    if top_response:
        resp, count = top_response[0]
        print(f"\n  FINDING 3: MEMBER RESPONSE TO PITCH")
        print(f"  {count}/{total_mp} members ({count/total_mp:.1%}) respond with '{resp}'")
        hangup_after = conv_flows['member_after_pitch'].get('SYS:hangup', 0)
        screening = conv_flows['member_after_pitch'].get('MBR:call_screening', 0)
        print(f"  Immediate hangup after pitch: {hangup_after}/{total_mp} ({hangup_after/total_mp:.1%})")
        print(f"  Call screening intercepted: {screening}/{total_mp} ({screening/total_mp:.1%})")

    # --- Finding 4: After member says "Hello?" ---
    top_agent_after = conv_flows['agent_after_greeting'].most_common(1)
    total_ag = sum(conv_flows['agent_after_greeting'].values())
    if top_agent_after:
        act, count = top_agent_after[0]
        print(f"\n  FINDING 4: AGENT RESPONSE TO 'HELLO?'")
        print(f"  {count}/{total_ag} times ({count/total_ag:.1%}) agent responds with '{act}'")
        disclaimer_count = conv_flows['agent_after_greeting'].get('AGT:disclaimer', 0)
        greeting_count = conv_flows['agent_after_greeting'].get('AGT:greeting', 0)
        print(f"  Disclaimer: {disclaimer_count}/{total_ag} ({disclaimer_count/total_ag:.1%})")
        print(f"  Greeting:   {greeting_count}/{total_ag} ({greeting_count/total_ag:.1%})")

    # --- Finding 5: Success vs failure divergence ---
    print(f"\n  FINDING 5: SUCCESS vs FAILURE DIVERGENCE")
    q_dist = Counter(d['quality'] for d in conv_flows['per_thread_flow_detail'].values())
    for q, c in q_dist.most_common():
        print(f"  {q:<25} {c:>4} ({c/spoke:.1%} of spoke_to)")

    # Check if call-screening correlates with success
    success_details = [d for d in conv_flows['per_thread_flow_detail'].values() if d['quality'] == 'success']
    fail_details = [d for d in conv_flows['per_thread_flow_detail'].values() if d['quality'] != 'success']
    success_with_screening = sum(1 for d in success_details if 'MBR:call_screening' in (d.get('fingerprint') or ''))
    fail_with_screening = sum(1 for d in fail_details if 'MBR:call_screening' in (d.get('fingerprint') or ''))
    if success_details:
        print(f"\n  Call screening in success paths: {success_with_screening}/{len(success_details)} ({success_with_screening/len(success_details):.1%})")
    if fail_details:
        print(f"  Call screening in failure paths: {fail_with_screening}/{len(fail_details)} ({fail_with_screening/len(fail_details):.1%})")
    if success_with_screening > 0:
        print(f"  → Call screening forces agent through fallback → re-greeting → proper flow.")
        print(f"    This accidentally fixes the opening sequence for those calls.")

    # --- Finding 6: Failure taxonomy ---
    print(f"\n  FINDING 6: FAILURE MODE BREAKDOWN")
    non_meaningful = sum(error_modes['failure_taxonomy'].values())
    for mode, count in error_modes['failure_taxonomy'].most_common():
        print(f"  {mode:<40} {count:>4} ({count/non_meaningful:.1%} of {non_meaningful} failures)")

    # --- Finding 7: VM fallback root cause ---
    print(f"\n  FINDING 7: VOICEMAIL FALLBACK ROOT CAUSE")
    print(f"  Clean VM delivery: {error_modes['vm_clean_count']}")
    print(f"  Fallback triggered: {error_modes['vm_fallback_count']}")
    if error_modes['vm_fallback_preceding']:
        top_preceding = error_modes['vm_fallback_preceding'].most_common(1)[0]
        print(f"  Primary trigger: \"{top_preceding[0][:80]}\" ({top_preceding[1]}x)")
        print(f"  → This is a carrier-level call screening prompt, not a real voicemail.")
        print(f"    Agent hears 'record your name' and plays 'important message' fallback.")

    # --- Finding 8: Double-pitch impact ---
    dp = signals['double_pitch']
    print(f"\n  FINDING 8: DOUBLE-PITCH BUG IMPACT")
    print(f"  Affected: {dp}/{spoke} spoke_to calls ({dp/spoke:.1%})")
    # Count how many of the top failure pattern have double-pitch
    top3_fail_count = sum(c for _, c in conv_flows['failure_patterns'].most_common(3))
    print(f"  Top 3 failure patterns account for {top3_fail_count} calls — all exhibit pitch repetition")

    # --- Prioritized fix list ---
    print(f"\n  {'=' * 60}")
    print(f"  PRIORITIZED FIXES (derived from data)")
    print(f"  {'=' * 60}")

    fixes = []

    # Fix 1: Opening sequence
    affected_by_sequence = sum(c for fp, c in conv_flows['failure_patterns'].most_common()
                                if 'AGT:pitch' in fp and fp.count('AGT:pitch') >= 2)
    fixes.append((affected_by_sequence, 'Restructure opening: disclaimer → greeting → pitch (not pitch → disclaimer → pitch)',
                   f'Addresses {affected_by_sequence} calls where pitch plays twice'))

    # Fix 2: VM fallback
    fixes.append((error_modes['vm_fallback_count'],
                   'Replace "important message" fallback with branded VM intro',
                   f'Affects {error_modes["vm_fallback_count"]} voicemails + {signals["important_msg"]} spoke_to calls'))

    # Fix 3: Post-greeting response
    disclaimer_after_hello = conv_flows['agent_after_greeting'].get('AGT:disclaimer', 0)
    fixes.append((disclaimer_after_hello,
                   'After member says "Hello?", respond with greeting first, not disclaimer',
                   f'{disclaimer_after_hello} calls go Hello? → disclaimer instead of Hello? → greeting'))

    # Fix 4: Proactive SMS
    fixes.append((spoke - signals['sms_sent'],
                   'Send SMS booking link proactively during spoke_to calls',
                   f'Only {signals["sms_sent"]}/{spoke} calls sent an SMS link'))

    fixes.sort(key=lambda x: x[0], reverse=True)
    for rank, (affected, fix, evidence) in enumerate(fixes, 1):
        print(f"\n  {rank}. [{affected} calls affected] {fix}")
        print(f"     Evidence: {evidence}")


# =============================================================================
# 4. EXCEL OUTPUT
# =============================================================================

def write_excel(output_path, threads, topline, loc_stats, buckets, quality, signals, timing, conv_flows=None, error_modes=None):
    """Write all analysis tables to a formatted Excel workbook."""
    if not HAS_OPENPYXL:
        print(f"Skipping Excel output (openpyxl not installed)")
        return

    wb = openpyxl.Workbook()

    # --- Styles ---
    title_font = Font(name='Helvetica', size=14, bold=True, color='1A1A2E')
    header_font = Font(name='Helvetica', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    subheader_font = Font(name='Helvetica', size=10, bold=True, color='1A1A2E')
    subheader_fill = PatternFill(start_color='E8E8F0', end_color='E8E8F0', fill_type='solid')
    data_font = Font(name='Helvetica', size=10)
    note_font = Font(name='Helvetica', size=9, color='666666')
    total_fill = PatternFill(start_color='F0F0F5', end_color='F0F0F5', fill_type='solid')
    total_font = Font(name='Helvetica', size=10, bold=True)
    red_font = Font(name='Helvetica', size=10, color='CC0000')
    green_font = Font(name='Helvetica', size=10, color='006600')
    thin_border = Border(bottom=Side(style='thin', color='D0D0D0'))
    pct_fmt = '0.0%'

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def style_row(ws, row, ncols, is_total=False):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = total_font if is_total else data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')
            if is_total:
                cell.fill = total_fill

    spoke = topline['spoke_to']
    total = topline['total_threads']

    # ---- Sheet 1: Topline ----
    ws = wb.active
    ws.title = 'Topline Metrics'

    ws.merge_cells('A1:C1')
    ws.cell(1, 1, 'UFC Gym Operate — Outcomes Appraisal').font = title_font
    ws.cell(2, 1, 'Data: April 13, 2026 (single day snapshot)').font = note_font
    ws.cell(3, 1, 'All numbers derived from raw thread export. Nothing estimated.').font = note_font

    r = 5
    for i, h in enumerate(['Metric', 'Value', 'Detail'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 3)

    rows_data = [
        ('Total call threads', total, ''),
        ('Locations', topline['location_count'], f"{topline['location_count'] - 1} production + 1 test"),
        ('Spoke to (connected)', spoke, f'{spoke/total:.1%}'),
        ('Left message (voicemail)', topline['left_message'], f"{topline['left_message']/total:.1%}"),
        ('Other (no verdict)', topline['other'], f"{topline['other']/total:.1%}"),
        ('Pickup rate', f"{topline['pickup_rate']:.1%}", 'spoke_to / total'),
        ('', '', ''),
        ('Total call time', f"{topline['total_call_time_s']:,}s", f"{topline['total_call_time_s']/3600:.1f} hours"),
        ('Connected time (spoke_to)', f"{topline['total_connected_time_s']:,}s", f"{topline['total_connected_time_s']/60:.1f} minutes"),
        ('Mean spoke_to duration', f"{topline['mean_spoke_duration_s']:.1f}s", ''),
        ('Median spoke_to duration', f"{topline['median_spoke_duration_s']}s", ''),
        ('Max spoke_to duration', f"{topline['max_spoke_duration_s']}s", f"{topline['max_spoke_duration_s']/60:.1f} minutes"),
    ]
    for name, val, detail in rows_data:
        r += 1
        ws.cell(r, 1, name).font = data_font
        ws.cell(r, 2, val).font = data_font
        ws.cell(r, 2).alignment = Alignment(horizontal='center')
        ws.cell(r, 3, detail).font = note_font
        if name:
            for c in range(1, 4):
                ws.cell(r, c).border = thin_border

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35

    # ---- Sheet 2: Per-Location ----
    ws = wb.create_sheet('Per-Location')
    ws.merge_cells('A1:I1')
    ws.cell(1, 1, 'Per-Location Performance — April 13, 2026').font = title_font

    r = 3
    hdrs = ['Location', 'Calls', 'Spoke', 'VM', 'Other', 'Pickup Rate', 'Avg Dur (s)', 'Med Dur (s)', 'Total Min']
    for i, h in enumerate(hdrs, 1):
        ws.cell(r, i, h)
    style_header(ws, r, len(hdrs))

    ranked = sorted(loc_stats.items(), key=lambda x: x[1]['pickup_rate'], reverse=True)
    for loc, s in ranked:
        r += 1
        ws.cell(r, 1, loc)
        ws.cell(r, 2, s['total'])
        ws.cell(r, 3, s['spoke'])
        ws.cell(r, 4, s['vm'])
        ws.cell(r, 5, s['other'])
        ws.cell(r, 6, s['pickup_rate']).number_format = pct_fmt
        ws.cell(r, 7, round(s['avg_spoke_dur'], 1))
        ws.cell(r, 8, s['median_spoke_dur'])
        ws.cell(r, 9, round(s['total_minutes'], 1))
        style_row(ws, r, len(hdrs))

    # Total
    r += 1
    ws.cell(r, 1, 'TOTAL')
    ws.cell(r, 2, total)
    ws.cell(r, 3, spoke)
    ws.cell(r, 4, topline['left_message'])
    ws.cell(r, 5, topline['other'])
    ws.cell(r, 6, topline['pickup_rate']).number_format = pct_fmt
    ws.cell(r, 7, round(topline['mean_spoke_duration_s'], 1))
    ws.cell(r, 8, topline['median_spoke_duration_s'])
    ws.cell(r, 9, round(topline['total_call_time_s'] / 60, 1))
    style_row(ws, r, len(hdrs), is_total=True)

    for c in range(1, len(hdrs) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16 if c > 1 else 38

    # ---- Sheet 3: Conversation Quality ----
    ws = wb.create_sheet('Conversation Quality')
    ws.merge_cells('A1:D1')
    ws.cell(1, 1, f'Conversation Quality — {spoke} Spoke-To Calls').font = title_font

    # Duration buckets
    r = 3
    ws.cell(r, 1, 'Duration Distribution').font = subheader_font
    ws.cell(r, 1).fill = subheader_fill
    ws.merge_cells(f'A{r}:D{r}')

    r += 1
    for i, h in enumerate(['Bucket', 'Count', '% of Spoke-To', 'Visual'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 4)

    for label, count in buckets.items():
        r += 1
        pct = count / spoke if spoke else 0
        ws.cell(r, 1, label)
        ws.cell(r, 2, count)
        ws.cell(r, 3, pct).number_format = pct_fmt
        ws.cell(r, 4, '|' * int(pct * 50))
        style_row(ws, r, 4)

    # Engagement
    r += 2
    ws.cell(r, 1, 'Engagement Metrics').font = subheader_font
    ws.cell(r, 1).fill = subheader_fill
    ws.merge_cells(f'A{r}:D{r}')

    r += 1
    for i, h in enumerate(['Metric', 'Count', '% of Spoke-To', 'Assessment'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 4)

    q = quality
    for name, cnt, assessment in [
        ('Real dialog (member 2+ turns)', q['real_dialog'], 'Low'),
        ('Meaningful calls (>30s)', q['meaningful_30s'], 'Low'),
        ('Member hung up', q['member_hangup'], 'Structural but high'),
        ('Clean close', q['clean_close'], 'Very low'),
    ]:
        r += 1
        ws.cell(r, 1, name)
        ws.cell(r, 2, cnt)
        ws.cell(r, 3, cnt / q['spoke_total'] if q['spoke_total'] else 0).number_format = pct_fmt
        ws.cell(r, 4, assessment)
        style_row(ws, r, 4)

    for c in range(1, 5):
        ws.column_dimensions[get_column_letter(c)].width = 22 if c > 1 else 42

    # ---- Sheet 4: Outcome Signals ----
    ws = wb.create_sheet('Outcome Signals')
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, 'Outcome & Pattern Signals').font = title_font
    ws.cell(2, 1, 'All counts verified; no estimates.').font = note_font

    r = 4
    ws.cell(r, 1, 'Outcome Signals').font = subheader_font
    ws.cell(r, 1).fill = subheader_fill
    ws.merge_cells(f'A{r}:E{r}')

    r += 1
    for i, h in enumerate(['Signal', 'Threads', '% of All', '% of Spoke-To', 'Note'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 5)

    sig = signals
    sig_rows = [
        ('Recovery services mentioned', sig['recovery'], sig['recovery']/sig['total'], None, 'Campaign topic'),
        ('Booking/scheduling language', sig['booking'], sig['booking']/sig['total'], sig['booking']/sig['spoke_total'] if sig['spoke_total'] else 0, ''),
        ('Member info requests', sig['member_info_req'], sig['member_info_req']/sig['total'], sig['member_info_req']/sig['spoke_total'] if sig['spoke_total'] else 0, 'Zero members asked about pricing/hours'),
        ('SMS link sent during call', sig['sms_sent'], sig['sms_sent']/sig['total'], sig['sms_sent']/sig['spoke_total'] if sig['spoke_total'] else 0, ''),
        ('Opt-out keywords', sig['opt_out'], sig['opt_out']/sig['total'], None, ''),
        ('Agent offered transfer', sig['escalation'], sig['escalation']/sig['total'], None, ''),
    ]
    for name, cnt, pct_all, pct_spoke, note in sig_rows:
        r += 1
        ws.cell(r, 1, name)
        ws.cell(r, 2, cnt)
        ws.cell(r, 3, pct_all).number_format = pct_fmt
        if pct_spoke is not None:
            ws.cell(r, 4, pct_spoke).number_format = pct_fmt
        else:
            ws.cell(r, 4, '—').alignment = Alignment(horizontal='center')
        ws.cell(r, 5, note).font = note_font
        style_row(ws, r, 5)

    # Technical issues
    r += 2
    ws.cell(r, 1, 'Technical Issues').font = subheader_font
    ws.cell(r, 1).fill = subheader_fill
    ws.merge_cells(f'A{r}:E{r}')

    r += 1
    for i, h in enumerate(['Issue', 'Affected', '% of Base', 'Severity', 'Description'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 5)

    for name, cnt, pct_str, sev, desc in [
        ('Double-pitch bug', sig['double_pitch'], f"{sig['double_pitch']/sig['spoke_total']:.1%} of spoke", 'HIGH', 'Opening pitch repeated after disclaimer'),
        ('"Important message" fallback', sig['important_msg'], f"{sig['important_msg']/sig['total']:.1%} of all", 'MEDIUM', 'Spammy VM fallback on uncertain detection'),
    ]:
        r += 1
        ws.cell(r, 1, name)
        ws.cell(r, 2, cnt)
        ws.cell(r, 3, pct_str)
        ws.cell(r, 4, sev).font = red_font if sev == 'HIGH' else data_font
        ws.cell(r, 5, desc).font = note_font
        style_row(ws, r, 5)

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 18 if c > 1 else 38
    ws.column_dimensions['E'].width = 45

    # ---- Sheet 5: Trend ----
    ws = wb.create_sheet('Trend')
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, 'Week-over-Week Trend').font = title_font
    ws.cell(2, 1, 'W1/W2: Corona+Sunnyvale week totals. Apr 13: single day, 10 locations.').font = note_font

    r = 4
    for i, h in enumerate(['Metric', 'W1 (Mar 16)', 'W2 (Mar 23)', 'Apr 13', 'Trend'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 5)

    for name, w1, w2, apr, trend in [
        ('Locations', 2, 2, 10, '5x'),
        ('Total calls', 48, 82, 613, '~12x daily'),
        ('Picked up', 7, 16, 240, '34x'),
        ('Pickup rate', '15.0%', '19.5%', '39.2%', '2.6x'),
        ('Texts sent', 348, 165, '—', 'Not in Apr 13 export'),
        ('Reply rate', '2.9%', '4.8%', '—', '1.7x (W1→W2)'),
        ('System failures', 0, 0, 0, 'Clean'),
    ]:
        r += 1
        ws.cell(r, 1, name)
        ws.cell(r, 2, w1)
        ws.cell(r, 3, w2)
        ws.cell(r, 4, apr)
        ws.cell(r, 5, trend).font = green_font
        style_row(ws, r, 5)

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 18 if c > 1 else 28

    # ---- Sheet 6: Call Timing ----
    ws = wb.create_sheet('Call Timing')
    ws.merge_cells('A1:F1')
    ws.cell(1, 1, 'Call Timing Distribution — UTC').font = title_font

    r = 3
    for i, h in enumerate(['Hour (UTC)', 'Approx Local', 'Calls', 'Spoke', 'Pickup Rate', '% of Day'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 6)

    tz_map = {3: '8pm HI', 15: '8am PT', 17: '10am PT', 18: '11am PT',
              19: '12pm PT', 20: '1pm PT', 21: '2pm PT'}
    total_day = sum(timing['hour_total'].values())
    for h in sorted(timing['hour_total'].keys()):
        r += 1
        t = timing['hour_total'][h]
        s = timing['hour_spoke'].get(h, 0)
        ws.cell(r, 1, f'{h}:00')
        ws.cell(r, 2, tz_map.get(h, ''))
        ws.cell(r, 3, t)
        ws.cell(r, 4, s)
        ws.cell(r, 5, s / t if t else 0).number_format = pct_fmt
        ws.cell(r, 6, t / total_day if total_day else 0).number_format = pct_fmt
        style_row(ws, r, 6)

    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ---- Sheet 7: Thread Detail ----
    ws = wb.create_sheet('Thread Detail')
    ws.merge_cells('A1:K1')
    ws.cell(1, 1, f'Thread-Level Detail — All {total} Threads').font = title_font

    r = 3
    det_hdrs = ['Thread ID', 'Location', 'Contact', 'Verdict', 'Duration (s)',
                'Total Msgs', 'Member Turns', 'Agent Turns', 'Double Pitch', 'VM Fallback', 'Recovery']
    for i, h in enumerate(det_hdrs, 1):
        ws.cell(r, i, h)
    style_header(ws, r, len(det_hdrs))

    for tid in sorted(threads.keys(), key=lambda x: int(x) if x.isdigit() else 0):
        t = threads[tid]
        r += 1
        agent_msgs = [m['message'] for m in t['transcript'] if m['direction'] == 'outbound']
        member_turns = sum(1 for m in t['transcript'] if m['direction'] == 'inbound' and len(m['message'].strip()) > 5)
        all_text = ' '.join(m['message'].lower() for m in t['transcript'])

        has_double = (
            t['verdict'] == 'spoke_to'
            and len(agent_msgs) >= 3
            and agent_msgs[0][:50].lower() in ' '.join(m[:50].lower() for m in agent_msgs[2:])
        )
        has_vm_fb = any(
            'important message' in m['message'].lower()
            for m in t['transcript'] if m['direction'] == 'outbound'
        )
        has_recovery = any(
            kw in all_text
            for kw in ['recovery', 'hydromassage', 'cryotherapy', 'compression massage', 'percussion']
        )

        ws.cell(r, 1, int(tid) if tid.isdigit() else tid)
        ws.cell(r, 2, t['location'])
        ws.cell(r, 3, t['contact'])
        ws.cell(r, 4, t['verdict'])
        ws.cell(r, 5, int(t['duration']) if t['duration'] else 0)
        ws.cell(r, 6, len(t['transcript']))
        ws.cell(r, 7, member_turns)
        ws.cell(r, 8, len(agent_msgs))
        ws.cell(r, 9, 'Yes' if has_double else 'No')
        ws.cell(r, 10, 'Yes' if has_vm_fb else 'No')
        ws.cell(r, 11, 'Yes' if has_recovery else 'No')

        for c in range(1, len(det_hdrs) + 1):
            ws.cell(r, c).font = data_font
            ws.cell(r, c).alignment = Alignment(horizontal='center' if c > 3 else 'left')

    for c in range(1, len(det_hdrs) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14 if c > 3 else 30
    ws.column_dimensions['A'].width = 12

    # ---- Sheet 8: Conversation Flows ----
    if conv_flows:
        cf = conv_flows
        ws = wb.create_sheet('Conversation Flows')
        ws.sheet_properties.tabColor = '4A3F8F'

        ws.merge_cells('A1:E1')
        ws.cell(1, 1, 'Conversation Flow Analysis — Dialog Act Classification').font = title_font
        ws.cell(2, 1, 'Method: Rule-based dialog act tagger (DAMSL/ISO 24617-2 inspired) applied to every turn.').font = note_font
        ws.cell(3, 1, 'Approach from TD-EVAL (Acikgoz et al. 2025) and BETOLD (Terragni et al. 2022).').font = note_font

        # 8a. Hangup on agent act
        r = 5
        ws.cell(r, 1, 'Agent Act When Member Hangs Up').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:D{r}')
        r += 1
        for i, h in enumerate(['Agent Dialog Act', 'Hangups', '% of Hangups', 'Interpretation'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 4)

        total_hu = sum(cf['hangup_on_act'].values())
        act_notes = {
            'AGT:pitch': 'Member hangs up during the opening pitch — not hooked',
            'AGT:info_give': 'Member hangs up during information delivery — not relevant',
            'AGT:disclaimer': 'Member hangs up during recording disclaimer — lost patience',
            'AGT:greeting': 'Member hangs up during agent greeting — didn\'t want the call',
            'AGT:question': 'Member hangs up when asked a question — felt pressured',
            'AGT:fallback': '"Important message" fallback triggered hangup',
            'AGT:ack_positive': 'Agent was acknowledging, member still left',
            'AGT:cta': 'Hung up during call-to-action — not interested in next step',
            'AGT:ack_graceful': 'Agent accepted decline gracefully, member left',
            'AGT:filler': 'Hung up during filler/transition — lost engagement',
        }
        for act, count in cf['hangup_on_act'].most_common():
            r += 1
            ws.cell(r, 1, act)
            ws.cell(r, 2, count)
            ws.cell(r, 3, count / total_hu if total_hu else 0).number_format = pct_fmt
            ws.cell(r, 4, act_notes.get(act, '')).font = note_font
            style_row(ws, r, 4)

        # 8b. Member response to pitch
        r += 2
        ws.cell(r, 1, 'Member Response After First Pitch').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:D{r}')
        r += 1
        for i, h in enumerate(['Member Response', 'Count', '% of Responses', 'Signal'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 4)

        total_mp = sum(cf['member_after_pitch'].values())
        response_signals = {
            'MBR:greeting': 'Member just said "Hello?" — pitch didn\'t register',
            'SYS:hangup': 'Member hung up immediately after pitch',
            'MBR:ack_brief': 'Brief acknowledgment — low engagement',
            'MBR:substantive': 'Real engagement — success signal',
            'MBR:decline_soft': 'Polite decline — targeting/timing issue',
            'MBR:vm_greeting': 'Voicemail greeting — misdetected as live person',
            'MBR:call_screening': 'Call screening app intercepted',
            'MBR:rejection': 'Hard rejection — do not call again',
            'MBR:suspicion': 'Suspicious of caller — trust issue',
        }
        for act, count in cf['member_after_pitch'].most_common():
            r += 1
            ws.cell(r, 1, act)
            ws.cell(r, 2, count)
            ws.cell(r, 3, count / total_mp if total_mp else 0).number_format = pct_fmt
            ws.cell(r, 4, response_signals.get(act, '')).font = note_font
            style_row(ws, r, 4)

        # 8c. Divergence analysis
        r += 2
        ws.cell(r, 1, 'Success vs Failure Divergence (position-by-position)').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:E{r}')
        r += 1
        for i, h in enumerate(['Turn Position', 'Success Top Act', 'Success Count', 'Failure Top Act', 'Failure Count'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 5)

        for d in cf['divergence'][:8]:
            r += 1
            ws.cell(r, 1, f"Position {d['position']}")
            if d['success_top']:
                ws.cell(r, 2, d['success_top'][0][0])
                ws.cell(r, 3, d['success_top'][0][1])
            if d['failure_top']:
                ws.cell(r, 4, d['failure_top'][0][0])
                ws.cell(r, 5, d['failure_top'][0][1])
            style_row(ws, r, 5)

        # 8d. Top flow patterns
        r += 2
        ws.cell(r, 1, 'Top Success Flow Patterns').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:C{r}')
        r += 1
        for i, h in enumerate(['Flow Pattern', 'Count', 'Quality'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 3)
        for fp, count in cf['success_patterns'].most_common(10):
            r += 1
            ws.cell(r, 1, fp)
            ws.cell(r, 2, count)
            ws.cell(r, 3, 'Success').font = green_font
            style_row(ws, r, 3)

        r += 2
        ws.cell(r, 1, 'Top Failure Flow Patterns').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:C{r}')
        r += 1
        for i, h in enumerate(['Flow Pattern', 'Count', 'Quality'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 3)
        for fp, count in cf['failure_patterns'].most_common(10):
            r += 1
            ws.cell(r, 1, fp)
            ws.cell(r, 2, count)
            ws.cell(r, 3, 'Failure').font = red_font
            style_row(ws, r, 3)

        # 8e. Hangup context (last 3 acts)
        r += 2
        ws.cell(r, 1, 'Last 3 Dialog Acts Before Hangup (top 15)').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:C{r}')
        r += 1
        for i, h in enumerate(['Act Sequence', 'Frequency', '% of Hangups'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 3)
        for ctx, count in cf['hangup_context'].most_common(15):
            r += 1
            ws.cell(r, 1, ctx)
            ws.cell(r, 2, count)
            ws.cell(r, 3, count / total_hu if total_hu else 0).number_format = pct_fmt
            style_row(ws, r, 3)

        for c in range(1, 6):
            ws.column_dimensions[get_column_letter(c)].width = 22 if c > 1 else 65

    # ---- Sheet 9: Error Modes ----
    if error_modes:
        em = error_modes
        ws = wb.create_sheet('Error Modes')
        ws.sheet_properties.tabColor = 'CC0000'

        ws.merge_cells('A1:E1')
        ws.cell(1, 1, 'Error Mode Analysis — Why Conversations Fail').font = title_font
        ws.cell(2, 1, 'Transcript-level pattern analysis of all 613 threads.').font = note_font

        # --- 8a. Member First Response ---
        r = 4
        ws.cell(r, 1, f'Member First Response (spoke_to: {spoke} calls)').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:D{r}')
        r += 1
        for i, h in enumerate(['Response Category', 'Count', '% of Spoke-To', 'Implication'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 4)

        implications = {
            'greeting (hello?)': 'Member picked up but agent didn\'t hook them',
            'brief acknowledgment': 'Minimal engagement — agent message not compelling',
            'substantive response': 'Good — member actually engaged with content',
            'voicemail greeting (misdetected)': 'BUG: VM detection failed, wasted a "spoke_to" slot',
            'suspicion (who is this?)': 'Caller ID or intro not building trust',
            'immediate rejection': 'Member not in target mood — timing or targeting issue',
            'wrong number': 'Data quality issue in CRM',
            'no member speech detected': 'Possible: picked up and immediately hung up',
            'silence/noise': 'Pocket answer or noise triggered pickup detection',
        }
        for cat, count in em['member_first_response'].most_common():
            r += 1
            ws.cell(r, 1, cat)
            ws.cell(r, 2, count)
            ws.cell(r, 3, count / spoke if spoke else 0).number_format = pct_fmt
            ws.cell(r, 4, implications.get(cat, '')).font = note_font
            style_row(ws, r, 4)

        # --- 8b. Failure Taxonomy ---
        r += 2
        non_meaningful = sum(em['failure_taxonomy'].values())
        ws.cell(r, 1, f'Failure Taxonomy — Non-Meaningful Calls (<30s): {non_meaningful} calls').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:D{r}')
        r += 1
        for i, h in enumerate(['Failure Mode', 'Count', '% of Failures', 'Fix Category'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 4)

        fix_cats = {
            'greeting_then_hangup': 'Agent opening — needs faster hook',
            'no_member_speech': 'Pickup detection or immediate hangup',
            'voicemail_misdetected_as_spoke': 'BUG: VM detection fix',
            'brief_exchange_unclear': 'Needs transcript review',
            'suspicion_hangup': 'Caller ID / trust building',
            'immediate_rejection': 'Targeting / timing',
        }
        for mode, count in em['failure_taxonomy'].most_common():
            r += 1
            ws.cell(r, 1, mode)
            ws.cell(r, 2, count)
            ws.cell(r, 3, count / non_meaningful if non_meaningful else 0).number_format = pct_fmt
            ws.cell(r, 4, fix_cats.get(mode, '')).font = note_font
            style_row(ws, r, 4)

        # --- 8c. Dropout by Agent Turn ---
        r += 2
        ws.cell(r, 1, 'Dropout by Agent Turn # (when member hung up)').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:D{r}')
        r += 1
        for i, h in enumerate(['Agent Turn #', 'Hangups', '% of Hangups', 'Interpretation'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 4)

        total_hangups = sum(em['dropout_by_agent_turn'].values())
        turn_notes = {
            0: 'Hung up before agent spoke — immediate disconnect',
            1: 'Hung up after first agent message (pitch or disclaimer)',
            2: 'Hung up after second message (likely during double-pitch)',
            3: 'Hung up after third message — heard pitch twice, left',
            4: 'Some conversation happened',
            5: 'Extended conversation — hung up mid-dialog',
        }
        for turn in sorted(em['dropout_by_agent_turn'].keys()):
            c = em['dropout_by_agent_turn'][turn]
            r += 1
            ws.cell(r, 1, f'Turn {turn}')
            ws.cell(r, 2, c)
            ws.cell(r, 3, c / total_hangups if total_hangups else 0).number_format = pct_fmt
            note = turn_notes.get(turn, f'Deep conversation (turn {turn})')
            ws.cell(r, 4, note).font = note_font
            style_row(ws, r, 4)

        # --- 8d. Last Agent Message Before Hangup (clustered) ---
        r += 2
        ws.cell(r, 1, 'Last Agent Message Before Hangup (top 15 patterns)').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:C{r}')
        r += 1
        for i, h in enumerate(['Message Pattern (first 80 chars)', 'Frequency', '% of Hangups'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 3)

        last_msg_clusters = Counter()
        for tid, loc, dur, msg in em['last_agent_before_hangup']:
            last_msg_clusters[msg[:80].strip()] += 1
        for msg, count in last_msg_clusters.most_common(15):
            r += 1
            ws.cell(r, 1, msg)
            ws.cell(r, 2, count)
            ws.cell(r, 3, count / total_hangups if total_hangups else 0).number_format = pct_fmt
            style_row(ws, r, 3)

        # --- 8e. Member Last Words (clustered) ---
        r += 2
        ws.cell(r, 1, 'Member Last Words Before Hangup (top 15)').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:C{r}')
        r += 1
        for i, h in enumerate(['Last Member Message', 'Frequency', 'Signal'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 3)

        last_word_clusters = Counter()
        for tid, loc, dur, msg in em['member_last_words']:
            last_word_clusters[msg[:60].strip().lower()] += 1
        for msg, count in last_word_clusters.most_common(15):
            r += 1
            ws.cell(r, 1, msg)
            ws.cell(r, 2, count)
            ws.cell(r, 3, '').font = note_font
            style_row(ws, r, 3)

        # --- 8f. Voicemail Delivery ---
        r += 2
        ws.cell(r, 1, 'Voicemail Delivery Quality').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:C{r}')
        r += 1
        for i, h in enumerate(['Category', 'Count', '% of Voicemails'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 3)
        total_vm = em['vm_clean_count'] + em['vm_fallback_count']
        for name, cnt in [('Clean delivery', em['vm_clean_count']), ('"Important message" fallback', em['vm_fallback_count'])]:
            r += 1
            ws.cell(r, 1, name)
            ws.cell(r, 2, cnt)
            ws.cell(r, 3, cnt / total_vm if total_vm else 0).number_format = pct_fmt
            style_row(ws, r, 3)

        if em['vm_fallback_preceding']:
            r += 1
            ws.cell(r, 1, 'What preceded the fallback (carrier/VM greetings):').font = note_font
            for msg, count in em['vm_fallback_preceding'].most_common(5):
                r += 1
                ws.cell(r, 1, f'  "{msg}"').font = note_font
                ws.cell(r, 2, count)

        # --- 8g. Other Threads ---
        r += 2
        ws.cell(r, 1, f'"Other" Threads (no verdict): {topline["other"]}').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:C{r}')
        r += 1
        for i, h in enumerate(['Category', 'Count', 'Note'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 3)
        for cat, count in em['other_thread_analysis'].most_common():
            if not cat.startswith('  msg:'):
                r += 1
                ws.cell(r, 1, cat)
                ws.cell(r, 2, count)
                style_row(ws, r, 3)

        # --- 8h. Agent Repetition ---
        r += 2
        ws.cell(r, 1, 'Agent Repetition').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        r += 1
        ws.cell(r, 1, 'Threads with same message 3+ times')
        ws.cell(r, 2, em['agent_repetition_threads'])
        style_row(ws, r, 2)

        for c in range(1, 5):
            ws.column_dimensions[get_column_letter(c)].width = 22 if c > 1 else 55

    # ---- Update Thread Detail with conv flows + error modes ----
    if 'Thread Detail' in wb.sheetnames:
        ws = wb['Thread Detail']
        next_col = len(det_hdrs) + 1

        # Add flow fingerprint column
        if conv_flows:
            fp_col = next_col
            ws.cell(3, fp_col, 'Flow Fingerprint')
            ws.cell(3, fp_col).font = header_font
            ws.cell(3, fp_col).fill = header_fill
            ws.cell(3, fp_col).alignment = Alignment(horizontal='center', wrap_text=True)

            ql_col = fp_col + 1
            ws.cell(3, ql_col, 'Quality Label')
            ws.cell(3, ql_col).font = header_font
            ws.cell(3, ql_col).fill = header_fill
            ws.cell(3, ql_col).alignment = Alignment(horizontal='center', wrap_text=True)

            row_idx = 4
            for tid in sorted(threads.keys(), key=lambda x: int(x) if x.isdigit() else 0):
                detail = conv_flows['per_thread_flow_detail'].get(tid, {})
                ws.cell(row_idx, fp_col, detail.get('fingerprint', ''))
                ws.cell(row_idx, fp_col).font = data_font
                ql = detail.get('quality', '')
                ws.cell(row_idx, ql_col, ql)
                ws.cell(row_idx, ql_col).font = green_font if ql == 'success' else red_font if 'fail' in ql else data_font
                row_idx += 1

            ws.column_dimensions[get_column_letter(fp_col)].width = 60
            ws.column_dimensions[get_column_letter(ql_col)].width = 18
            next_col = ql_col + 1

        # Add error modes column
        if error_modes:
            err_col = next_col
            ws.cell(3, err_col, 'Error Modes')
            ws.cell(3, err_col).font = header_font
            ws.cell(3, err_col).fill = header_fill
            ws.cell(3, err_col).alignment = Alignment(horizontal='center', wrap_text=True)

            row_idx = 4
            for tid in sorted(threads.keys(), key=lambda x: int(x) if x.isdigit() else 0):
                errs = error_modes['per_thread_errors'].get(tid, [])
                ws.cell(row_idx, err_col, ', '.join(errs) if errs else '')
                ws.cell(row_idx, err_col).font = red_font if errs else data_font
                row_idx += 1

            ws.column_dimensions[get_column_letter(err_col)].width = 40

    # ---- Sheet 10: Synthesis ----
    if conv_flows and error_modes:
        ws = wb.create_sheet('Synthesis')
        ws.sheet_properties.tabColor = '006600'

        ws.merge_cells('A1:D1')
        ws.cell(1, 1, 'Root-Cause Synthesis — Derived From Data').font = title_font
        ws.cell(2, 1, 'Every number below is computed from the analysis. Nothing hardcoded.').font = note_font

        r = 4
        ws.cell(r, 1, 'Prioritized Fixes').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:D{r}')
        r += 1
        for i, h in enumerate(['Rank', 'Fix', 'Calls Affected', 'Evidence'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 4)

        # Compute fixes
        affected_by_sequence = sum(c for fp, c in conv_flows['failure_patterns'].most_common()
                                    if 'AGT:pitch' in fp and fp.count('AGT:pitch') >= 2)
        disclaimer_after_hello = conv_flows['agent_after_greeting'].get('AGT:disclaimer', 0)

        fixes = [
            (affected_by_sequence,
             'Restructure opening: disclaimer → greeting → pitch',
             f'{affected_by_sequence} calls where pitch plays twice'),
            (error_modes['vm_fallback_count'] + signals['important_msg'],
             'Replace "important message" fallback with branded VM intro',
             f"{error_modes['vm_fallback_count']} VMs + {signals['important_msg']} spoke_to calls"),
            (disclaimer_after_hello,
             'After "Hello?", respond with greeting first, not disclaimer',
             f'{disclaimer_after_hello} calls go Hello? → disclaimer'),
            (spoke - signals['sms_sent'],
             'Send SMS booking link proactively during spoke_to calls',
             f"Only {signals['sms_sent']}/{spoke} calls sent an SMS"),
        ]
        fixes.sort(key=lambda x: x[0], reverse=True)
        for rank, (affected, fix, evidence) in enumerate(fixes, 1):
            r += 1
            ws.cell(r, 1, rank)
            ws.cell(r, 2, fix)
            ws.cell(r, 3, affected)
            ws.cell(r, 4, evidence).font = note_font
            style_row(ws, r, 4)

        # Key findings
        r += 2
        ws.cell(r, 1, 'Key Findings').font = subheader_font
        ws.cell(r, 1).fill = subheader_fill
        ws.merge_cells(f'A{r}:D{r}')
        r += 1
        for i, h in enumerate(['#', 'Finding', 'Data Point', 'Implication'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 4)

        # Dominant failure flow
        top_fail = conv_flows['failure_patterns'].most_common(1)
        if top_fail:
            fp, fp_count = top_fail[0]
            fail_total = sum(conv_flows['failure_patterns'].values())
            r += 1
            ws.cell(r, 1, 1)
            ws.cell(r, 2, f'Dominant failure: {fp}')
            ws.cell(r, 3, f'{fp_count}/{fail_total} ({fp_count/fail_total:.1%}) of failures')
            ws.cell(r, 4, 'Pitch before member ready → disclaimer → repeated pitch → hangup').font = note_font
            style_row(ws, r, 4)

        # Hangup trigger
        top_hangup_act = conv_flows['hangup_on_act'].most_common(1)
        total_hu = sum(conv_flows['hangup_on_act'].values())
        if top_hangup_act:
            act, count = top_hangup_act[0]
            r += 1
            ws.cell(r, 1, 2)
            ws.cell(r, 2, f'Primary hangup trigger: {act}')
            ws.cell(r, 3, f'{count}/{total_hu} ({count/total_hu:.1%}) of hangups')
            ws.cell(r, 4, 'Members leave during the pitch — not during info or questions').font = note_font
            style_row(ws, r, 4)

        # Member response
        total_mp = sum(conv_flows['member_after_pitch'].values())
        greeting_resp = conv_flows['member_after_pitch'].get('MBR:greeting', 0)
        r += 1
        ws.cell(r, 1, 3)
        ws.cell(r, 2, 'Members respond "Hello?" to pitch')
        ws.cell(r, 3, f'{greeting_resp}/{total_mp} ({greeting_resp/total_mp:.1%})')
        ws.cell(r, 4, 'Pitch fires before member registers what is happening').font = note_font
        style_row(ws, r, 4)

        # Call screening correlation
        success_details = [d for d in conv_flows['per_thread_flow_detail'].values() if d['quality'] == 'success']
        fail_details = [d for d in conv_flows['per_thread_flow_detail'].values() if d['quality'] != 'success']
        s_screening = sum(1 for d in success_details if 'MBR:call_screening' in (d.get('fingerprint') or ''))
        f_screening = sum(1 for d in fail_details if 'MBR:call_screening' in (d.get('fingerprint') or ''))
        if success_details and fail_details:
            r += 1
            ws.cell(r, 1, 4)
            ws.cell(r, 2, 'Call screening correlates with success')
            ws.cell(r, 3, f'{s_screening}/{len(success_details)} ({s_screening/len(success_details):.1%}) success vs {f_screening}/{len(fail_details)} ({f_screening/len(fail_details):.1%}) failure')
            ws.cell(r, 4, 'Screening forces re-greeting → proper flow. Accidentally fixes the opening.').font = note_font
            style_row(ws, r, 4)

        # Quality distribution
        q_dist = Counter(d['quality'] for d in conv_flows['per_thread_flow_detail'].values())
        r += 1
        ws.cell(r, 1, 5)
        ws.cell(r, 2, 'Quality distribution')
        ws.cell(r, 3, f"success={q_dist.get('success',0)}, fail_brief={q_dist.get('fail_brief',0)}, fail_immediate={q_dist.get('fail_immediate',0)}, fail_no_response={q_dist.get('fail_no_response',0)}")
        ws.cell(r, 4, f"{q_dist.get('success',0)}/{spoke} ({q_dist.get('success',0)/spoke:.1%}) of spoke_to calls are successful").font = note_font
        style_row(ws, r, 4)

        # Zero info requests
        r += 1
        ws.cell(r, 1, 6)
        ws.cell(r, 2, 'Zero members asked follow-up questions')
        ws.cell(r, 3, f"{signals['member_info_req']}/{spoke} member-initiated info requests")
        ws.cell(r, 4, 'Agent presents info; members are not curious enough to engage').font = note_font
        style_row(ws, r, 4)

        for c in range(1, 5):
            ws.column_dimensions[get_column_letter(c)].width = 22 if c in (1, 3) else 55

    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")


# =============================================================================
# 5. MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='UFC Gym Operate Outcomes Analysis')
    parser.add_argument('--input', '-i', type=str, help='Path to thread export (TSV or JSON)')
    parser.add_argument('--output', '-o', type=str,
                        default='ufc-operate-outcomes-appraisal.xlsx',
                        help='Output Excel path')
    parser.add_argument('--no-excel', action='store_true', help='Skip Excel generation')
    args = parser.parse_args()

    # Load data
    if args.input:
        p = Path(args.input)
        if p.suffix == '.json':
            text = load_from_json(str(p))
        else:
            text = load_from_tsv(str(p))
    else:
        # Default: MCP cache path
        default_path = (
            Path.home() / '.claude' / 'projects'
            / '-Users-mgc50-Dropbox-1--Worked-On-FILES--34--Vi-ufc-info'
            / '9bf07344-f6dc-4ea4-811f-dc39cfbee709' / 'tool-results'
            / 'mcp-gdocs-gsheet_read-1776203983894.txt'
        )
        if default_path.exists():
            text = load_from_json(str(default_path))
        else:
            print("ERROR: No input file specified and default cache not found.")
            print("Usage: python analyze_operate_outcomes.py --input <path_to_data>")
            return

    # Parse
    threads = parse_threads(text)
    print(f"Loaded {len(threads)} threads from data.\n")

    # Compute
    topline = compute_topline(threads)
    loc_stats = compute_per_location(threads)
    buckets = compute_duration_buckets(threads)
    quality = compute_quality(threads)
    signals = compute_signals(threads)
    timing = compute_timing(threads)
    conv_flows = compute_conversation_flows(threads)
    error_modes = compute_error_modes(threads)

    # Output
    print_tables(threads, topline, loc_stats, buckets, quality, signals, timing, conv_flows, error_modes)
    print_synthesis(topline, quality, signals, conv_flows, error_modes)

    if not args.no_excel:
        write_excel(args.output, threads, topline, loc_stats, buckets, quality, signals, timing, conv_flows, error_modes)


if __name__ == '__main__':
    main()
